import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import threading
import subprocess
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from PIL import Image, ImageTk
    _PIL_OK = True
except ImportError:
    _PIL_OK = False

# ─────────────────────── Amazon Invoice 处理常量 ───────────────────────
# Charge Type = "Shipping Charge" 下的 Charge Code → 输出列名，顺序为 H~N
CHARGE_CODE_MAP = {
    "Base charge":                                  "Base charge（基础运费）",
    "High Cube Surcharge":                          "High Cube Surcharge（体积附加费）",
    "Delivery Area Surcharge":                      "Delivery Area Surcharge（偏远地区）",
    "Misdeclaration Handling Charge - Dimensions":  "Misdeclaration Handling Charge - Dimensions（错误申报费）",
    "Base Rate Adjustment":                         "Base Rate Adjustment（费率调整）",
    "Non-Conveyable Surcharge":                     "Non-Conveyable Surcharge（不可传送附加费）",
    "Additional Handling Fees: Girth":              "Additional Handling Fees: Girth（附加处理费：围长）",
}
CHARGE_COLS_OUTPUT = list(CHARGE_CODE_MAP.values())   # H~N

# O 列：Charge Type = "Shipping Charge Adjustment" 的汇总
ADJ_COL = "Shipping Charge Adjustment-运费调整"        # O

OUTPUT_COLUMNS = [
    "Tracking ID",                  # A
    "To Postcode",                  # B
    "Reference",                    # C
    "Billable weight (kg)",         # D
    "Length (cm)",                  # E
    "Width (cm)",                   # F
    "Height (cm)",                  # G
    *CHARGE_COLS_OUTPUT,            # H~N
    ADJ_COL,                        # O  Shipping Charge Adjustment
    "合计（不含税）",                # P = SUM(H:O)
    "发票金额（含税）",              # Q = L4 含税总额
    "备注",                         # R
]

# ─────────────────────── 颜色 & 字体常量 ───────────────────────
BG_COLOR      = "#F5F6FA"   # 浅色主背景
SIDEBAR_COLOR = "#FFFFFF"   # 白色侧边栏
ACCENT        = "#2D5BE3"   # 深蓝按钮
ACCENT_HOVER  = "#1E46C7"
DANGER        = "#C0392B"   # 深红删除按钮
DANGER_HOVER  = "#962D22"
TEXT_DARK     = "#1A1D2E"   # 深色文字
TEXT_GRAY     = "#7A82A8"
BORDER_COLOR  = "#E4E7F0"
SUCCESS_COLOR = "#1A7A50"   # 深绿合并按钮
SUCCESS_HOVER = "#136040"

FONT_TITLE  = ("Microsoft YaHei UI", 18, "bold")
FONT_BTN    = ("Microsoft YaHei UI", 10, "bold")
FONT_NORMAL = ("Microsoft YaHei UI", 10)
FONT_SMALL  = ("Microsoft YaHei UI", 9)


# ─────────────────────── 扁平按钮 ───────────────────────
class FlatButton(tk.Button):
    def __init__(self, parent, text, command, bg=ACCENT, fg="white",
                 hover_bg=ACCENT_HOVER, width=130, height=36,
                 font=FONT_BTN, **kwargs):
        super().__init__(
            parent, text=text, command=command,
            bg=bg, fg=fg, activebackground=hover_bg, activeforeground=fg,
            font=font, relief="flat", bd=0, cursor="hand2",
            width=int(width // 8), pady=6,
            **kwargs
        )
        self._bg       = bg
        self._hover_bg = hover_bg
        self._cmd      = command
        self.bind("<Enter>", lambda e: self.config(bg=hover_bg))
        self.bind("<Leave>", lambda e: self.config(bg=bg))

    def set_state(self, enabled: bool):
        if enabled:
            self.config(bg=SUCCESS_COLOR, state="normal", cursor="hand2")
            self.bind("<Enter>", lambda e: self.config(bg=SUCCESS_HOVER))
            self.bind("<Leave>", lambda e: self.config(bg=SUCCESS_COLOR))
        else:
            self.config(bg="#AAAAAA", state="disabled", cursor="arrow")
            self.unbind("<Enter>")
            self.unbind("<Leave>")


# ─────────────────────── 主应用 ───────────────────────
class ExcelMergerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Amazon Invoice 合并工具")
        self.geometry("940x660")
        self.minsize(740, 520)
        self.configure(bg=BG_COLOR)
        self._files: list[str] = []
        self._mascot_img = None   # 防止 GC
        self._build_ui()
        self._center_window()

    def _center_window(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"+{(sw-w)//2}+{(sh-h)//2}")

    # ─── 界面构建 ───
    def _build_ui(self):
        left = tk.Frame(self, bg=SIDEBAR_COLOR, width=270)
        left.pack(side="left", fill="y")
        left.pack_propagate(False)
        self._build_left_panel(left)

        tk.Frame(self, bg=BORDER_COLOR, width=1).pack(side="left", fill="y")

        right = tk.Frame(self, bg=BG_COLOR)
        right.pack(side="left", fill="both", expand=True)
        self._build_right_panel(right)

    def _build_left_panel(self, parent):
        # ── 顶部标题栏 ──
        header = tk.Frame(parent, bg="#2D5BE3", height=64)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="📦  Invoice 合并", font=FONT_TITLE,
                 bg="#2D5BE3", fg="white").pack(expand=True)

        # ── 说明文字 ──
        desc = tk.Frame(parent, bg=SIDEBAR_COLOR)
        desc.pack(fill="x", padx=18, pady=(16, 6))
        tk.Label(desc, text="使用说明", font=("Microsoft YaHei UI", 10, "bold"),
                 bg=SIDEBAR_COLOR, fg=TEXT_DARK, anchor="w").pack(fill="x")
        tips = (
            "① 导入一个或多个 Amazon Invoice\n"
            "   (.xlsx / .xls 格式)\n"
            "② 自动按 Tracking ID 透视汇总\n"
            "③ 生成含税 / 不含税合计列\n"
            "④ 点击「开始合并」导出总表"
        )
        tk.Label(desc, text=tips, font=FONT_SMALL, bg=SIDEBAR_COLOR,
                 fg=TEXT_GRAY, justify="left", wraplength=220).pack(fill="x", pady=(4, 0))

        tk.Frame(parent, bg=BORDER_COLOR, height=1).pack(fill="x", padx=18, pady=10)

        # ── 操作按钮 ──
        btn_frame = tk.Frame(parent, bg=SIDEBAR_COLOR)
        btn_frame.pack(fill="x", padx=18)

        FlatButton(btn_frame, "＋ 添加文件", self._add_files,
                   bg="#2D5BE3", hover_bg="#1E46C7", width=220, height=38).pack(fill="x", pady=3)
        FlatButton(btn_frame, "✕  删除选中", self._remove_selected,
                   bg=DANGER, hover_bg=DANGER_HOVER, width=220, height=38).pack(fill="x", pady=3)
        FlatButton(btn_frame, "清空列表", self._clear_all,
                   bg="#555C7A", hover_bg="#404660", width=220, height=38).pack(fill="x", pady=3)

        tk.Frame(parent, bg=BORDER_COLOR, height=1).pack(fill="x", padx=18, pady=10)

        self._merge_btn = FlatButton(
            parent, "▶  开始合并", self._start_merge,
            bg=SUCCESS_COLOR, hover_bg=SUCCESS_HOVER, width=220, height=44,
            font=("Microsoft YaHei UI", 11, "bold")
        )
        self._merge_btn.pack(padx=18, pady=2, fill="x")

        # ── 底部吉祥物图片 ──
        self._load_mascot(parent)

        tk.Label(parent, text="v2.0  |  Amazon Invoice Merger",
                 font=FONT_SMALL, bg=SIDEBAR_COLOR, fg=TEXT_GRAY).pack(
            side="bottom", pady=6)

    def _load_mascot(self, parent):
        # 查找图片：优先同目录，其次打包资源目录
        candidates = [
            os.path.join(os.path.dirname(os.path.abspath(__file__)), "mascot.png"),
            os.path.join(os.path.dirname(sys.executable), "mascot.png"),
        ]
        img_path = next((p for p in candidates if os.path.exists(p)), None)
        if not img_path or not _PIL_OK:
            return
        try:
            img = Image.open(img_path)
            # 缩放到宽 230，保持比例
            w_target = 230
            ratio = w_target / img.width
            h_target = int(img.height * ratio)
            img = img.resize((w_target, h_target), Image.LANCZOS)
            self._mascot_img = ImageTk.PhotoImage(img)
            lbl = tk.Label(parent, image=self._mascot_img,
                           bg=SIDEBAR_COLOR, cursor="hand2")
            lbl.pack(side="bottom", pady=(0, 2))
        except Exception:
            pass

    def _build_right_panel(self, parent):
        top = tk.Frame(parent, bg=BG_COLOR)
        top.pack(fill="x", padx=24, pady=(20, 0))
        self._file_count_var = tk.StringVar(value="已导入 0 个文件")
        tk.Label(top, textvariable=self._file_count_var,
                 font=("Microsoft YaHei UI", 12, "bold"),
                 bg=BG_COLOR, fg=TEXT_DARK).pack(side="left")

        list_frame = tk.Frame(parent, bg=BG_COLOR)
        list_frame.pack(fill="both", expand=True, padx=24, pady=12)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Custom.Treeview",
                        background=SIDEBAR_COLOR, fieldbackground=SIDEBAR_COLOR,
                        foreground=TEXT_DARK, rowheight=32, font=FONT_NORMAL, borderwidth=0)
        style.configure("Custom.Treeview.Heading",
                        background=BORDER_COLOR, foreground=TEXT_GRAY,
                        font=("Microsoft YaHei UI", 9, "bold"), relief="flat")
        style.map("Custom.Treeview",
                  background=[("selected", "#EBF2FF")],
                  foreground=[("selected", ACCENT)])

        columns = ("序号", "文件名", "路径", "大小")
        self._tree = ttk.Treeview(list_frame, columns=columns,
                                   show="headings", selectmode="extended",
                                   style="Custom.Treeview")
        self._tree.heading("序号",  text="#")
        self._tree.heading("文件名", text="文件名")
        self._tree.heading("路径",  text="文件路径")
        self._tree.heading("大小",  text="大小")
        self._tree.column("序号",  width=45,  anchor="center", stretch=False)
        self._tree.column("文件名", width=200, anchor="w")
        self._tree.column("路径",  width=360, anchor="w")
        self._tree.column("大小",  width=80,  anchor="center", stretch=False)

        vsb = ttk.Scrollbar(list_frame, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        self._tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        status_bar = tk.Frame(parent, bg=SIDEBAR_COLOR, height=48)
        status_bar.pack(fill="x", side="bottom")
        status_bar.pack_propagate(False)
        tk.Frame(status_bar, bg=BORDER_COLOR, height=1).pack(fill="x")

        inner = tk.Frame(status_bar, bg=SIDEBAR_COLOR)
        inner.pack(fill="both", expand=True, padx=24)

        self._progress = ttk.Progressbar(inner, mode="indeterminate", length=160)
        self._progress.pack(side="right", pady=12)

        self._status_var = tk.StringVar(value="就绪，请导入 Amazon Invoice 文件")
        tk.Label(inner, textvariable=self._status_var,
                 font=FONT_SMALL, bg=SIDEBAR_COLOR, fg=TEXT_GRAY).pack(side="left", pady=12)

    # ─── 文件操作 ───
    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title="选择 Amazon Invoice 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls *.XLSX *.XLS"), ("所有文件", "*.*")]
        )
        added = 0
        for p in paths:
            if p not in self._files:
                self._files.append(p)
                added += 1
        self._refresh_list()
        if added:
            self._set_status(f"已添加 {added} 个文件")

    def _remove_selected(self):
        selected = self._tree.selection()
        if not selected:
            messagebox.showinfo("提示", "请先在列表中选中要删除的文件。")
            return
        indices = sorted([self._tree.index(s) for s in selected], reverse=True)
        for i in indices:
            self._files.pop(i)
        self._refresh_list()
        self._set_status(f"已删除 {len(selected)} 个文件")

    def _clear_all(self):
        if not self._files:
            return
        if messagebox.askyesno("确认", "确定要清空文件列表吗？"):
            self._files.clear()
            self._refresh_list()
            self._set_status("文件列表已清空")

    def _refresh_list(self):
        self._tree.delete(*self._tree.get_children())
        for i, path in enumerate(self._files, 1):
            name = os.path.basename(path)
            size = self._fmt_size(os.path.getsize(path))
            self._tree.insert("", "end", values=(i, name, path, size))
        self._file_count_var.set(f"已导入 {len(self._files)} 个文件")

    @staticmethod
    def _fmt_size(b: int) -> str:
        if b < 1024:       return f"{b} B"
        if b < 1024 ** 2:  return f"{b/1024:.1f} KB"
        return f"{b/1024**2:.1f} MB"

    def _set_status(self, msg: str):
        self._status_var.set(msg)

    # ─── 合并入口 ───
    def _start_merge(self):
        if not self._files:
            messagebox.showwarning("提示", "请先添加至少一个 Excel 文件。")
            return
        save_path = filedialog.asksaveasfilename(
            title="保存合并总表",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile="Amazon_合并总表.xlsx"
        )
        if not save_path:
            return
        self._set_status("正在处理，请稍候…")
        self._progress.start(10)
        self._merge_btn.set_state(False)
        threading.Thread(target=self._do_merge, args=(save_path,), daemon=True).start()

    # ─── 核心处理逻辑（后台线程） ───
    def _do_merge(self, save_path: str):
        try:
            all_rows: list[dict] = []
            for path in self._files:
                self.after(0, self._set_status, f"处理：{os.path.basename(path)}")
                rows = self._process_invoice_file(path)
                all_rows.extend(rows)

            # 跨文件按 Invoice 号排序
            all_rows.sort(key=lambda r: r.get("_invoice_number", ""))

            self.after(0, self._set_status, "写入文件…")
            self._write_output(all_rows, save_path)
            self.after(0, self._on_done, save_path, len(all_rows))
        except Exception as exc:
            import traceback
            self.after(0, self._on_error, traceback.format_exc())

    def _process_invoice_file(self, path: str) -> list[dict]:
        """
        解析单个 Amazon Invoice XLSX：
        - 自动定位表头行（含 'Tracking ID' 的行）
        - 按 (Invoice号, Tracking ID) 分组，相同 TID 跨发票不合并
        - Charge Type = Shipping Charge  → 透视各 Charge Code 为 H~N 列
        - Charge Type = Shipping Charge Adjustment → 汇总到 O 列
        - 按 Invoice 号排序后返回
        """
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        all_rows_raw = list(ws.iter_rows(values_only=True))
        wb.close()

        # ── 1. 读取发票元数据（前4行） ──
        # G2 = Invoice number（第2行第7列）
        try:
            invoice_number = str(all_rows_raw[1][6] or "")
        except IndexError:
            invoice_number = ""

        # L4 = Total Tax-Inclusive Charge（第4行第12列）
        invoice_total_incl_tax = None
        try:
            raw_val = all_rows_raw[3][11]
            if raw_val is not None:
                invoice_total_incl_tax = float(str(raw_val).replace(",", ""))
        except (IndexError, ValueError, TypeError):
            invoice_total_incl_tax = None

        # ── 2. 定位表头行 ──
        header_row_idx = None
        col_map: dict[str, int] = {}
        for r_idx, row in enumerate(all_rows_raw, 1):
            if row and "Tracking ID" in row:
                header_row_idx = r_idx
                col_map = {str(v): i for i, v in enumerate(row) if v is not None}
                break
        if header_row_idx is None:
            raise ValueError(f"未在文件中找到表头行（含 'Tracking ID'）：\n{path}")

        # ── 3. 读取数据行，按 (invoice_number, tid) 分组 ──
        def col(row_vals, name):
            idx = col_map.get(name)
            return row_vals[idx] if idx is not None and idx < len(row_vals) else None

        # key: (invoice_number, tid) → 汇总字典
        records: dict[tuple, dict] = {}

        for row in all_rows_raw[header_row_idx:]:
            tid = col(row, "Tracking ID")
            if not tid:
                continue

            charge_code  = col(row, "Charge Code")
            charge_type  = col(row, "Charge Type")
            tax_excl_val = col(row, "Tax Exclusive Charge Value (GBP)")

            # 跳过汇总行（Charge Code 为空 = Total 小计行）
            if not charge_code:
                continue

            ctype_str = str(charge_type).strip() if charge_type else ""
            key = (invoice_number, str(tid))

            if key not in records:
                records[key] = {
                    "_invoice_number":    invoice_number,
                    "Tracking ID":        tid,
                    "To Postcode":        col(row, "To Postcode"),
                    "Reference":          col(row, "Reference"),
                    "Billable weight (kg)": col(row, "Billable weight (kg)"),
                    "Length (cm)":        col(row, "Length (cm)"),
                    "Width (cm)":         col(row, "Width (cm)"),
                    "Height (cm)":        col(row, "Height (cm)"),
                    **{v: None for v in CHARGE_CODE_MAP.values()},
                    ADJ_COL: None,
                }

            val = float(tax_excl_val) if tax_excl_val is not None else 0.0
            code_str = str(charge_code).strip() if charge_code else ""

            if ctype_str == "Shipping Charge Adjustment":
                # 优先按 Charge Code 匹配到对应列（如 K=错误申报费、L=费率调整）
                # 匹配不到的才汇总进 O 列（运费调整）
                output_col = CHARGE_CODE_MAP.get(code_str)
                if output_col:
                    records[key][output_col] = round(
                        (records[key][output_col] or 0) + val, 4
                    )
                else:
                    records[key][ADJ_COL] = round((records[key][ADJ_COL] or 0) + val, 4)
            elif ctype_str in ("Shipping Charge", ""):
                # H~N 列：按 Charge Code 分配
                output_col = CHARGE_CODE_MAP.get(code_str)
                if output_col:
                    records[key][output_col] = round(
                        (records[key][output_col] or 0) + val, 4
                    )

        # ── 4. 计算合计列，注入发票元数据 ──
        result = []
        for key, rec in records.items():
            total_excl = sum(
                rec[c] for c in CHARGE_COLS_OUTPUT if rec[c] is not None
            )
            adj = rec[ADJ_COL] or 0
            rec["合计（不含税）"]  = round(total_excl + adj, 2)
            rec["发票金额（含税）"] = invoice_total_incl_tax
            rec["备注"] = None
            result.append(rec)

        # 按 Invoice 号排序（同一文件内已有序，跨文件合并后再统一排序）
        result.sort(key=lambda r: r["_invoice_number"])
        return result

    # ─── 写出 Excel ───
    def _write_output(self, rows: list[dict], path: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "合并总表"

        # ── 样式定义 ──
        thin = Side(style="thin", color="D0D5E8")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        hdr_fill  = PatternFill("solid", fgColor="4F8EF7")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        num_fmt_2 = '#,##0.00'    # 两位小数数值列
        num_fmt_4 = '#,##0.0000'  # 重量等四位精度列

        # 特殊背景：O 列（合计不含税）橙黄，P 列（含税）绿
        o_fill = PatternFill("solid", fgColor="FFF2CC")
        p_fill = PatternFill("solid", fgColor="E2EFDA")
        o_font = Font(name="Arial", bold=True, size=10, color="7F4F00")
        p_font = Font(name="Arial", bold=True, size=10, color="1D5E2A")

        # ── 写表头 ──
        ws.row_dimensions[1].height = 36
        for c_idx, col_name in enumerate(OUTPUT_COLUMNS, 1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = border

        # ── 写数据 ──
        adj_col_idx = OUTPUT_COLUMNS.index(ADJ_COL) + 1          # O 列
        o_col_idx   = OUTPUT_COLUMNS.index("合计（不含税）") + 1  # P 列
        p_col_idx   = OUTPUT_COLUMNS.index("发票金额（含税）") + 1 # Q 列

        for r_idx, rec in enumerate(rows, 2):
            is_even = (r_idx % 2 == 0)
            row_bg  = PatternFill("solid", fgColor="F7F9FF") if is_even else None

            for c_idx, col_name in enumerate(OUTPUT_COLUMNS, 1):
                value = rec.get(col_name)
                cell  = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border    = border
                cell.alignment = Alignment(vertical="center",
                                           horizontal="right" if isinstance(value, (int, float)) else "left")

                # 数值格式
                if col_name in ("Billable weight (kg)",):
                    cell.number_format = num_fmt_4
                elif c_idx in range(8, o_col_idx + 1):   # H~P 费用+合计列
                    if value is not None:
                        cell.number_format = num_fmt_2

                # O 列（Adjustment）浅紫色
                if c_idx == adj_col_idx:
                    if r_idx == 1:
                        pass  # 表头已统一处理
                    else:
                        cell.number_format = num_fmt_2
                        if row_bg:
                            cell.fill = PatternFill("solid", fgColor="EDE7F6")
                # P 列（合计不含税）橙黄
                elif c_idx == o_col_idx:
                    cell.fill   = o_fill
                    cell.font   = o_font
                    cell.number_format = num_fmt_2
                # Q 列（含税）绿
                elif c_idx == p_col_idx:
                    cell.fill   = p_fill
                    cell.font   = p_font
                    cell.number_format = num_fmt_2
                # 普通斑马纹
                elif row_bg and cell.fill.fgColor.rgb == "00000000":
                    cell.fill = row_bg

        # ── 自动列宽 ──
        for c_idx, col_name in enumerate(OUTPUT_COLUMNS, 1):
            col_letter = get_column_letter(c_idx)
            sample_vals = [
                len(str(ws.cell(row=r, column=c_idx).value or ""))
                for r in range(1, min(ws.max_row + 1, 200))
            ]
            best_width = max(sample_vals) + 3 if sample_vals else 12
            ws.column_dimensions[col_letter].width = min(best_width, 38)

        # ── 冻结首行 + 自动筛选 ──
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_COLUMNS))}1"

        wb.save(path)

    # ─── 回调 ───
    def _on_done(self, path: str, row_count: int):
        self._progress.stop()
        self._merge_btn.set_state(True)
        self._set_status(f"完成！共处理 {row_count} 条记录")
        ans = messagebox.askyesno(
            "合并成功",
            f"总表已生成，共 {row_count} 条 Tracking ID 记录。\n\n"
            f"文件路径：{path}\n\n是否打开所在文件夹？"
        )
        if ans:
            folder = os.path.dirname(path)
            if sys.platform == "win32":
                subprocess.Popen(["explorer", folder])
            elif sys.platform == "darwin":
                subprocess.Popen(["open", folder])
            else:
                subprocess.Popen(["xdg-open", folder])

    def _on_error(self, tb: str):
        self._progress.stop()
        self._merge_btn.set_state(True)
        self._set_status("处理失败，请检查文件格式")
        messagebox.showerror("处理失败", f"发生错误：\n\n{tb}")


# ─────────────────────── 入口 ───────────────────────
if __name__ == "__main__":
    app = ExcelMergerApp()
    app.mainloop()
